import streamlit as st
import io
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page config ────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Touché · Retail Invoices",
    page_icon="✂️",
    layout="centered",
)

# ── Styling ────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,400;0,600;1,400&family=Jost:wght@300;400;500&display=swap');

html, body, [class*="css"] {
    font-family: 'Jost', sans-serif;
}

/* Hide Streamlit branding */
#MainMenu, footer, header { visibility: hidden; }

/* Page background */
.stApp {
    background-color: #f7f5f2;
}

/* Main header */
.touche-header {
    text-align: center;
    padding: 2.5rem 0 1.5rem;
    border-bottom: 1px solid #d4cfc8;
    margin-bottom: 2rem;
}
.touche-header h1 {
    font-family: 'Cormorant Garamond', serif;
    font-size: 2.8rem;
    font-weight: 600;
    color: #1a1a2e;
    letter-spacing: 0.05em;
    margin: 0;
}
.touche-header p {
    font-family: 'Jost', sans-serif;
    font-weight: 300;
    font-size: 0.85rem;
    color: #7a7570;
    letter-spacing: 0.15em;
    text-transform: uppercase;
    margin: 0.4rem 0 0;
}

/* Section labels */
.section-label {
    font-family: 'Jost', sans-serif;
    font-weight: 500;
    font-size: 0.75rem;
    letter-spacing: 0.2em;
    text-transform: uppercase;
    color: #4a3060;
    margin-bottom: 0.4rem;
}

/* Summary table */
.summary-table {
    width: 100%;
    border-collapse: collapse;
    font-family: 'Jost', sans-serif;
    font-size: 0.9rem;
    margin-top: 1rem;
}
.summary-table th {
    background: #1a1a2e;
    color: #fff;
    padding: 0.6rem 1rem;
    text-align: left;
    font-weight: 400;
    letter-spacing: 0.08em;
    font-size: 0.78rem;
    text-transform: uppercase;
}
.summary-table th:not(:first-child) { text-align: right; }
.summary-table td {
    padding: 0.55rem 1rem;
    border-bottom: 1px solid #e8e4df;
    color: #1a1a2e;
}
.summary-table td:not(:first-child) { text-align: right; }
.summary-table tr:nth-child(even) td { background: #f0ecf7; }
.summary-table tr.total-row td {
    background: #4a3060;
    color: #fff;
    font-weight: 500;
    border-bottom: none;
}
.summary-table tr.total-row td:not(:first-child) { text-align: right; }
.warning-text {
    color: #c0392b;
    font-size: 0.8rem;
    font-style: italic;
}

/* Upload boxes */
.upload-hint {
    font-size: 0.8rem;
    color: #7a7570;
    margin-top: 0.2rem;
    font-style: italic;
}

/* Divider */
.thin-rule {
    border: none;
    border-top: 1px solid #d4cfc8;
    margin: 2rem 0;
}
</style>

<div class="touche-header">
    <h1>Touché Hairdressing</h1>
    <p>Retail Stock Invoice Generator</p>
</div>
""", unsafe_allow_html=True)


# ── Helpers ────────────────────────────────────────────────────────────

SKIP_TERMS = {
    'Inspired Hair Supplies', 'Salon Success', 'Client Deposit Lost',
    'Grand Total', 'Description',
}
NON_PRODUCT = {'Lost Deposit'}

FILL_DARK   = "1A1A2E"
FILL_MID    = "E8E0F0"
FILL_WHITE  = "FFFFFF"
FILL_PURPLE = "4A3060"
POUND       = '£#,##0.00'

thin  = Side(style='thin', color='BBBBBB')
thick = Side(style='medium', color='888888')
brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
brd_t = Border(left=thin, right=thin, top=thick, bottom=thick)


def opfill(hex_col):
    return PatternFill('solid', start_color=hex_col, fgColor=hex_col)


def xls_to_csv_rows(uploaded_file):
    """Read an uploaded XLS file and return rows as a list of string lists."""
    try:
        data = uploaded_file.read()
        wb = xlrd.open_workbook(file_contents=data)
        ws = wb.sheet_by_index(0)
        rows = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    row.append('')
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    # Preserve numeric precision as string
                    row.append(str(cell.value))
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    row.append(str(cell.value))
                else:
                    row.append(str(cell.value).strip())
            rows.append(row)
        return rows
    except Exception as e:
        return None


def _cell(row, idx):
    """Safely get a stripped string value by column index."""
    try:
        return str(row[idx]).strip()
    except IndexError:
        return ''


def _find_col(rows, target, search_rows=15):
    """Scan the first N rows to find the column index of a header value."""
    for row in rows[:search_rows]:
        for j, val in enumerate(row):
            if str(val).strip() == target:
                return j
    return None


def parse_sales(rows):
    """
    Parse retail sales rows dynamically using header detection.

    Detects 'Description' and 'Qty' column positions from the header row.
    Products have qty in the Qty column; supplier subtotals and stylist
    summary rows have qty one column to the right (SalonIQ layout).

    Returns dict: {stylist_name: [{'product': str, 'qty': int}, ...]}
    """
    SUPPLIER_SUBTOTALS = {'Inspired Hair Supplies', 'Salon Success', 'Client Deposit Lost'}

    name_col = _find_col(rows, 'Description') or 1
    qty_col  = _find_col(rows, 'Qty') or 6
    sub_col  = qty_col + 1  # subtotals and stylist rows use the next column

    # First pass: find stylist names.
    # Stylist rows have a value in sub_col but not qty_col,
    # and immediately follow a supplier subtotal row with the same pattern.
    stylist_names = set()
    prev_was_subtotal = False
    for row in rows:
        name    = _cell(row, name_col)
        qty_val = _cell(row, qty_col)
        sub_val = _cell(row, sub_col)
        if not name:
            continue
        is_sub_row = bool(sub_val and not qty_val)
        if name in SUPPLIER_SUBTOTALS and is_sub_row:
            prev_was_subtotal = True
        elif is_sub_row and name not in SKIP_TERMS and name not in NON_PRODUCT \
                and name != 'Grand Total':
            if prev_was_subtotal:
                stylist_names.add(name)
            prev_was_subtotal = False
        else:
            prev_was_subtotal = False

    # Second pass: collect products per stylist
    stylists, pending = {}, []
    for row in rows:
        name    = _cell(row, name_col)
        qty_val = _cell(row, qty_col)
        sub_val = _cell(row, sub_col)
        if not name:
            continue
        if name in stylist_names:
            stylists[name] = list(pending)
            pending = []
        elif name == 'Grand Total':
            pending = []
        elif qty_val and not sub_val and name not in SKIP_TERMS and name not in NON_PRODUCT:
            try:
                pending.append({'product': name, 'qty': int(float(qty_val))})
            except ValueError:
                pass

    return stylists


def parse_stock(rows):
    """
    Parse stock valuation rows dynamically using header detection.

    Detects 'Description' and 'Cost' column positions from the header row.

    Returns dict: {product_name: cost_price}
    """
    name_col = _find_col(rows, 'Description') or 1
    cost_col = _find_col(rows, 'Cost')

    if cost_col is None:
        return {}

    stock = {}
    for row in rows:
        name     = _cell(row, name_col)
        cost_str = _cell(row, cost_col)
        if not name or not cost_str:
            continue
        try:
            cost = float(cost_str)
            if cost > 0:
                stock[name] = cost
        except ValueError:
            pass
    return stock


def parse_period(rows):
    """
    Extract the report period from the sales CSV.
    SalonIQ puts it on the row after 'RETAIL SALES BY TEAM MEMBER',
    formatted as e.g. '15/02/26 to 15/03/26'.
    Returns a formatted string like '15 Feb 2026 – 15 Mar 2026', or None.
    """
    from datetime import datetime
    for i, row in enumerate(rows):
        v = [c.strip() for c in row if c.strip()]
        if any('RETAIL SALES BY TEAM MEMBER' in val for val in v):
            # Period is on the next non-empty row
            for next_row in rows[i + 1:]:
                nv = [c.strip() for c in next_row if c.strip()]
                if nv:
                    raw = nv[0]
                    # Expect "DD/MM/YY to DD/MM/YY"
                    if ' to ' in raw:
                        parts = raw.split(' to ')
                        try:
                            d1 = datetime.strptime(parts[0].strip(), '%d/%m/%y')
                            d2 = datetime.strptime(parts[1].strip(), '%d/%m/%y')
                            return f"{d1.strftime('%-d %b %Y')} – {d2.strftime('%-d %b %Y')}"
                        except ValueError:
                            return raw  # return as-is if parsing fails
                    break
    return None


def build_workbook(salon_name, period, stylists, stock):
    wb = Workbook()
    wb.remove(wb.active)
    ordered = sorted(stylists.keys())

    for items in stylists.values():
        for item in items:
            item['cost'] = stock.get(item['product'])

    # ── Per-stylist sheets ─────────────────────────────────────────────
    for stylist in ordered:
        items = stylists[stylist]
        ws = wb.create_sheet(title=stylist)

        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 7
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14

        for row_num, (value, merge, fnt, fll, align) in enumerate([
            (f"Touché Hairdressing {salon_name}", 'A1:F1',
             Font(name='Arial', bold=True, size=14, color='FFFFFF'), opfill(FILL_DARK),
             Alignment(horizontal='center', vertical='center')),
            (f"Retail Stock Invoice  |  {period}", 'A2:F2',
             Font(name='Arial', size=11, color='FFFFFF'), opfill(FILL_DARK),
             Alignment(horizontal='center', vertical='center')),
            (f"Stylist: {stylist}", 'A3:F3',
             Font(name='Arial', bold=True, size=12, color=FILL_DARK), opfill(FILL_MID),
             Alignment(horizontal='left', vertical='center', indent=1)),
        ], start=1):
            ws.merge_cells(merge)
            c = ws[merge.split(':')[0]]
            c.value = value; c.font = fnt; c.fill = fll; c.alignment = align
            ws.row_dimensions[row_num].height = 26 if row_num == 1 else 20

        headers = ['Product', 'Qty', 'Unit Cost', 'Cost (ex VAT)', 'VAT (20%)', 'Total (inc VAT)']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
            c.fill = opfill(FILL_PURPLE)
            c.alignment = Alignment(horizontal='left' if col == 1 else 'center',
                                    vertical='center', indent=1 if col == 1 else 0)
            c.border = brd
        ws.row_dimensions[4].height = 18

        if not items:
            ws.merge_cells('A5:F5')
            c = ws['A5']
            c.value = "No retail product sales this period"
            c.font = Font(name='Arial', italic=True, size=10, color='888888')
            c.fill = opfill(FILL_WHITE)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = brd
            ws.row_dimensions[5].height = 16
            total_row = 6
            ws.merge_cells(f'A{total_row}:C{total_row}')
            c = ws.cell(row=total_row, column=1, value="TOTAL")
            c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
            c.fill = opfill(FILL_PURPLE); c.border = brd_t
            c.alignment = Alignment(horizontal='right', vertical='center', indent=1)
            for col in range(4, 7):
                c = ws.cell(row=total_row, column=col, value=0)
                c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
                c.fill = opfill(FILL_PURPLE); c.border = brd_t
                c.number_format = POUND
                c.alignment = Alignment(horizontal='right', vertical='center')
            ws.row_dimensions[total_row].height = 20
            ws.freeze_panes = 'A5'
            continue

        data_start = 5
        for i, item in enumerate(items):
            r = data_start + i
            row_fill = opfill(FILL_MID) if i % 2 == 0 else opfill(FILL_WHITE)
            cost = item['cost']
            qty  = item['qty']
            font = Font(name='Arial', size=10, color=FILL_DARK)

            c = ws.cell(row=r, column=1, value=item['product'])
            c.font = font; c.fill = row_fill; c.border = brd
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

            c = ws.cell(row=r, column=2, value=qty)
            c.font = font; c.fill = row_fill; c.border = brd
            c.alignment = Alignment(horizontal='center', vertical='center')

            if cost is not None:
                c = ws.cell(row=r, column=3, value=cost)
                c.font = font; c.fill = row_fill; c.border = brd
                c.number_format = POUND
                c.alignment = Alignment(horizontal='right', vertical='center')

                c = ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
                c.font = font; c.fill = row_fill; c.border = brd
                c.number_format = POUND
                c.alignment = Alignment(horizontal='right', vertical='center')

                c = ws.cell(row=r, column=5, value=f"=D{r}*0.2")
                c.font = font; c.fill = row_fill; c.border = brd
                c.number_format = POUND
                c.alignment = Alignment(horizontal='right', vertical='center')

                c = ws.cell(row=r, column=6, value=f"=D{r}+E{r}")
                c.font = font; c.fill = row_fill; c.border = brd
                c.number_format = POUND
                c.alignment = Alignment(horizontal='right', vertical='center')
            else:
                for col in range(3, 7):
                    c = ws.cell(row=r, column=col, value="⚠ PRICE NOT IN STOCK FILE")
                    c.font = Font(name='Arial', size=10, bold=True, color='C0392B')
                    c.fill = opfill('FDECEA'); c.border = brd
                    c.alignment = Alignment(horizontal='center', vertical='center')

            ws.row_dimensions[r].height = 16

        last_data = data_start + len(items) - 1
        total_row = last_data + 1

        ws.merge_cells(f'A{total_row}:C{total_row}')
        c = ws.cell(row=total_row, column=1, value="TOTAL")
        c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        c.fill = opfill(FILL_PURPLE); c.border = brd_t
        c.alignment = Alignment(horizontal='right', vertical='center', indent=1)

        for col, col_l in [(4, 'D'), (5, 'E'), (6, 'F')]:
            c = ws.cell(row=total_row, column=col,
                        value=f"=SUM({col_l}{data_start}:{col_l}{last_data})")
            c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
            c.fill = opfill(FILL_PURPLE); c.border = brd_t
            c.number_format = POUND
            c.alignment = Alignment(horizontal='right', vertical='center')

        ws.row_dimensions[total_row].height = 20
        ws.freeze_panes = 'A5'

    # ── Summary sheet ──────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="Summary", index=0)
    for col, width in zip('ABCD', [22, 14, 14, 14]):
        ws_sum.column_dimensions[col].width = width

    ws_sum.merge_cells('A1:D1')
    ws_sum['A1'] = f"Touché Hairdressing {salon_name} — Retail Stock Summary"
    ws_sum['A1'].font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    ws_sum['A1'].fill = opfill(FILL_DARK)
    ws_sum['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 26

    ws_sum.merge_cells('A2:D2')
    ws_sum['A2'] = period
    ws_sum['A2'].font = Font(name='Arial', size=11, color='FFFFFF')
    ws_sum['A2'].fill = opfill(FILL_DARK)
    ws_sum['A2'].alignment = Alignment(horizontal='center')
    ws_sum.row_dimensions[2].height = 18

    for col, h in enumerate(['Stylist', 'Items Sold', 'Cost ex VAT', 'Total inc VAT'], 1):
        c = ws_sum.cell(row=3, column=col, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = opfill(FILL_PURPLE)
        c.alignment = Alignment(horizontal='left' if col == 1 else 'center',
                                vertical='center', indent=1 if col == 1 else 0)
        c.border = brd

    for i, stylist in enumerate(ordered):
        r = 4 + i
        items = stylists[stylist]
        n_items = sum(it['qty'] for it in items)
        row_fill = opfill(FILL_MID) if i % 2 == 0 else opfill(FILL_WHITE)

        c = ws_sum.cell(row=r, column=1, value=stylist)
        c.font = Font(name='Arial', bold=True, size=10, color=FILL_DARK)
        c.fill = row_fill; c.border = brd
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        c = ws_sum.cell(row=r, column=2, value=n_items)
        c.font = Font(name='Arial', size=10, color=FILL_DARK)
        c.fill = row_fill; c.border = brd
        c.alignment = Alignment(horizontal='center', vertical='center')

        if not items:
            for col in [3, 4]:
                c = ws_sum.cell(row=r, column=col, value="—")
                c.font = Font(name='Arial', size=10, color='888888')
                c.fill = row_fill; c.border = brd
                c.alignment = Alignment(horizontal='center', vertical='center')
        else:
            last_row_on_sheet = 4 + len(items) + 1
            sheet_ref = f"'{stylist}'"
            c = ws_sum.cell(row=r, column=3,
                            value=f"={sheet_ref}!D{last_row_on_sheet}")
            c.font = Font(name='Arial', size=10, color=FILL_DARK)
            c.fill = row_fill; c.border = brd; c.number_format = POUND
            c.alignment = Alignment(horizontal='right', vertical='center')

            c = ws_sum.cell(row=r, column=4,
                            value=f"={sheet_ref}!F{last_row_on_sheet}")
            c.font = Font(name='Arial', size=10, color=FILL_DARK)
            c.fill = row_fill; c.border = brd; c.number_format = POUND
            c.alignment = Alignment(horizontal='right', vertical='center')

    grand_r = 4 + len(ordered)
    ws_sum.merge_cells(f'A{grand_r}:B{grand_r}')
    c = ws_sum.cell(row=grand_r, column=1, value="GRAND TOTAL")
    c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill = opfill(FILL_PURPLE); c.border = brd_t
    c.alignment = Alignment(horizontal='right', vertical='center', indent=1)
    for col, col_l in [(3, 'C'), (4, 'D')]:
        c = ws_sum.cell(row=grand_r, column=col,
                        value=f"=SUM({col_l}4:{col_l}{grand_r - 1})")
        c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        c.fill = opfill(FILL_PURPLE); c.border = brd_t; c.number_format = POUND
        c.alignment = Alignment(horizontal='right', vertical='center')

    return wb


def workbook_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ─────────────────────────────────────────────────────────────────

st.markdown('<div class="section-label">Salon</div>', unsafe_allow_html=True)
salon = st.selectbox("Salon", ["Caterham", "Purley"], label_visibility="collapsed")

st.markdown('<hr class="thin-rule">', unsafe_allow_html=True)

col3, col4 = st.columns(2)

with col3:
    st.markdown('<div class="section-label">Retail Sales by Team Member (XLS)</div>',
                unsafe_allow_html=True)
    sales_file = st.file_uploader("Sales", type=["xls", "xlsx"],
                                  label_visibility="collapsed")
    st.markdown('<div class="upload-hint">Export from SalonIQ → Reports → Retail Sales by Team Member</div>',
                unsafe_allow_html=True)

with col4:
    st.markdown('<div class="section-label">Stock Valuation Report (XLS)</div>',
                unsafe_allow_html=True)
    stock_file = st.file_uploader("Stock", type=["xls", "xlsx"],
                                  label_visibility="collapsed")
    st.markdown('<div class="upload-hint">Export from SalonIQ → Reports → Stock Valuation</div>',
                unsafe_allow_html=True)

st.markdown('<hr class="thin-rule">', unsafe_allow_html=True)

if st.button("Generate Invoices", use_container_width=True, type="primary"):
    if not sales_file or not stock_file:
        st.error("Please upload both files before generating.")
    else:
        with st.spinner("Processing…"):
            sales_rows = xls_to_csv_rows(sales_file)
            stock_rows = xls_to_csv_rows(stock_file)

            if sales_rows is None or stock_rows is None:
                st.error("Could not read one of the uploaded files. Please check they are valid XLS exports from SalonIQ.")
                st.stop()

            period = parse_period(sales_rows)
            if not period:
                period = "Period unknown"

            stylists = parse_sales(sales_rows)
            stock    = parse_stock(stock_rows)

            if not stylists:
                st.error("No stylist data found in the sales file. Please check the export.")
                st.stop()

            # Match costs
            for items in stylists.values():
                for item in items:
                    item['cost'] = stock.get(item['product'])

            # Build workbook
            wb = build_workbook(salon, period, stylists, stock)
            xlsx_bytes = workbook_to_bytes(wb)

        # ── Summary preview ────────────────────────────────────────────
        st.markdown(f"### {salon} · {period}")

        unmatched_products = set()
        rows_html = ""
        grand_ex = grand_inc = 0

        for stylist in sorted(stylists.keys()):
            items = stylists[stylist]
            matched = [i for i in items if i['cost'] is not None]
            bad     = [i['product'] for i in items if i['cost'] is None]
            ex  = sum(i['cost'] * i['qty'] for i in matched)
            inc = ex * 1.2
            n   = sum(i['qty'] for i in items)
            grand_ex += ex; grand_inc += inc
            unmatched_products.update(bad)

            note = ""
            if not items:
                note = '<span class="warning-text">No retail sales</span>'
            elif bad:
                note = f'<span class="warning-text">⚠ No price: {", ".join(bad)}</span>'

            ex_str  = f"£{ex:,.2f}"  if items else "—"
            inc_str = f"£{inc:,.2f}" if items else "—"

            rows_html += f"""
            <tr>
                <td><strong>{stylist}</strong>{('<br>' + note) if note else ''}</td>
                <td>{n}</td>
                <td>{ex_str}</td>
                <td>{inc_str}</td>
            </tr>"""

        rows_html += f"""
        <tr class="total-row">
            <td>Grand Total</td>
            <td></td>
            <td>£{grand_ex:,.2f}</td>
            <td>£{grand_inc:,.2f}</td>
        </tr>"""

        st.markdown(f"""
        <table class="summary-table">
            <thead>
                <tr>
                    <th>Stylist</th>
                    <th>Items</th>
                    <th>Cost ex VAT</th>
                    <th>Total inc VAT</th>
                </tr>
            </thead>
            <tbody>{rows_html}</tbody>
        </table>
        """, unsafe_allow_html=True)

        if unmatched_products:
            st.warning(
                f"**{len(unmatched_products)} product(s) could not be matched to the stock file** "
                f"and are flagged in the spreadsheet:\n\n" +
                "\n".join(f"- {p}" for p in sorted(unmatched_products))
            )

        st.markdown("<br>", unsafe_allow_html=True)
        filename = f"Touche_{salon}_Retail_Invoices_{period.replace(' ', '_').replace('–','-')}.xlsx"
        st.download_button(
            label="⬇  Download Invoice Spreadsheet",
            data=xlsx_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
